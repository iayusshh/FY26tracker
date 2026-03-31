from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


BASE_BG = "1B1F2A"
CARD_BG = "232838"
HEADER_BG = "2D3348"
SPACER_BG = "0F172A"
BORDER = "334155"
PROFIT = "00D26A"
LOSS = "FF4D4D"
INPUT = "4E8EF7"
PURPLE = "A855F7"
AMBER = "F59E0B"
CYAN = "06B6D4"
TEXT = "F1F5F9"
MUTED = "94A3B8"

QUARTERS = ["Q1 (Apr-Jun)", "Q2 (Jul-Sep)", "Q3 (Oct-Dec)", "Q4 (Jan-Mar)"]
MONTHS_FULL = [
    "Apr 2025",
    "May 2025",
    "Jun 2025",
    "Jul 2025",
    "Aug 2025",
    "Sep 2025",
    "Oct 2025",
    "Nov 2025",
    "Dec 2025",
    "Jan 2026",
    "Feb 2026",
    "Mar 2026",
]
EXPENSE_CATS = [
    "Rent/Housing",
    "Food & Groceries",
    "Subscriptions",
    "Transport/Fuel",
    "Shopping/Clothing",
    "Health/Medical",
    "Education",
    "Entertainment",
    "Utilities",
    "Other",
]


@dataclass
class BuildContext:
    wb: Workbook
    ws: any
    row: int = 1
    max_col: int = 14
    user_name_cell: str = ""


def hex_color(value: str) -> colors.Color:
    value = value.lstrip("#")
    return colors.HexColor(f"#{value}")


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def set_cell(ws, cell_ref: str, value=None, *, font=None, fill_color=None, alignment=None, border=None, number_format=None):
    cell = ws[cell_ref]
    cell.value = value
    if font:
        cell.font = font
    if fill_color:
        cell.fill = fill(fill_color)
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell


def style_range(ws, start_row: int, end_row: int, start_col: int, end_col: int, *, fill_color=None, font=None, border=None, alignment=None):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            if fill_color:
                cell.fill = fill(fill_color)
            if font:
                cell.font = font
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment


def currency_formula(fmt="₹#,##0.00;[Red]-₹#,##0.00"):
    return fmt


def percent_formula():
    return "0.00%"


def add_spacer(ctx: BuildContext, height: float = 8):
    ws = ctx.ws
    ws.row_dimensions[ctx.row].height = height
    style_range(ws, ctx.row, ctx.row, 1, ctx.max_col, fill_color=SPACER_BG, border=Border())
    ctx.row += 1


def add_section_header(ctx: BuildContext, title: str, color: str):
    ws = ctx.ws
    ws.merge_cells(start_row=ctx.row, start_column=1, end_row=ctx.row, end_column=ctx.max_col)
    set_cell(
        ws,
        f"A{ctx.row}",
        title,
        font=Font(name="Calibri", size=14, bold=True, color=TEXT),
        fill_color=color,
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    ws.row_dimensions[ctx.row].height = 24
    style_range(ws, ctx.row, ctx.row, 1, ctx.max_col, fill_color=color, border=thin_border())
    ctx.row += 1


def thin_border():
    side = Side(border_style="thin", color=BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def data_font(color=TEXT, bold=False, size=11):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def center():
    return Alignment(horizontal="center", vertical="center")


def left():
    return Alignment(horizontal="left", vertical="center")


def right():
    return Alignment(horizontal="right", vertical="center")


def add_dashboard(ctx: BuildContext):
    ws = ctx.ws
    border = thin_border()

    ws.merge_cells(start_row=ctx.row, start_column=1, end_row=ctx.row, end_column=ctx.max_col)
    set_cell(
        ws,
        f"A{ctx.row}",
        "FY 2025-26 | CONSOLIDATED FINANCIAL TRACKER",
        font=Font(name="Calibri", size=16, bold=True, color=TEXT),
        fill_color=HEADER_BG,
        alignment=center(),
        border=border,
    )
    ws.row_dimensions[ctx.row].height = 26
    title_row = ctx.row
    ctx.row += 1

    set_cell(ws, f"A{ctx.row}", "User Name", font=data_font(TEXT, True), fill_color=CARD_BG, alignment=left(), border=border)
    set_cell(ws, f"B{ctx.row}", "Ayush", font=data_font(INPUT, True), fill_color=BASE_BG, alignment=left(), border=border)
    ctx.user_name_cell = f"B{ctx.row}"
    set_cell(ws, f"D{ctx.row}", "Prepared On", font=data_font(TEXT, True), fill_color=CARD_BG, alignment=left(), border=border)
    set_cell(ws, f"E{ctx.row}", "=TODAY()", font=data_font(TEXT), fill_color=BASE_BG, alignment=center(), border=border, number_format="dd-mmm-yyyy")
    set_cell(ws, f"G{ctx.row}", "Total Income", font=data_font(TEXT, True), fill_color=CARD_BG, alignment=left(), border=border)
    set_cell(ws, f"H{ctx.row}", "=SUM(F4:F9)", font=data_font(TEXT, True), fill_color=BASE_BG, alignment=right(), border=border, number_format=currency_formula())
    set_cell(ws, f"J{ctx.row}", "Total Expenses", font=data_font(TEXT, True), fill_color=CARD_BG, alignment=left(), border=border)
    set_cell(ws, f"K{ctx.row}", "=N156", font=data_font(TEXT, True), fill_color=BASE_BG, alignment=right(), border=border, number_format=currency_formula())
    set_cell(ws, f"M{ctx.row}", "Net Savings", font=data_font(TEXT, True), fill_color=CARD_BG, alignment=left(), border=border)
    set_cell(ws, f"N{ctx.row}", "=H2-K2", font=data_font(TEXT, True), fill_color=BASE_BG, alignment=right(), border=border, number_format=currency_formula())
    ctx.row += 1

    headers = ["Income Stream", *QUARTERS, "FY Total", "% of Total"]
    for col_idx, text in enumerate(headers, start=1):
        set_cell(ws, f"{get_column_letter(col_idx)}{ctx.row}", text, font=data_font(TEXT, True), fill_color=CARD_BG, alignment=center(), border=border)
    summary_header_row = ctx.row
    ctx.row += 1

    streams = [
        ("F&O Trading (You)", ["G27", "H27", "I27", "J27"], "K27"),
        ("F&O Trading (Family)", ["G44", "H44", "I44", "J44"], "K44"),
        ("Equity (You)", ["F48", "F49", "F50", "F51"], "F52"),
        ("Equity (Family)", ["F67", "F68", "F69", "F70"], "F71"),
        ("Crypto", ["D87", "0", "0", "0"], "D87"),
        ("Business/Freelance", ["N115", "N116", "N117", "N118"], "E112"),
    ]

    for idx, (label, qrefs, total_ref) in enumerate(streams):
        row_num = ctx.row
        row_fill = CARD_BG if idx % 2 == 0 else BASE_BG
        set_cell(ws, f"A{row_num}", label, font=data_font(TEXT), fill_color=row_fill, alignment=left(), border=border)
        for q_idx in range(4):
            value = f"={qrefs[q_idx]}" if qrefs[q_idx] != "0" else "=0"
            set_cell(ws, f"{get_column_letter(2 + q_idx)}{row_num}", value, font=data_font(TEXT), fill_color=row_fill, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"F{row_num}", f"={total_ref}", font=data_font(TEXT, True), fill_color=row_fill, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"G{row_num}", f"=IFERROR(F{row_num}/$H$2,0)", font=data_font(TEXT), fill_color=row_fill, alignment=right(), border=border, number_format=percent_formula())
        ctx.row += 1

    set_cell(ws, f"A{ctx.row}", "Net Position", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=left(), border=border)
    set_cell(ws, f"B{ctx.row}", "=H2", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=right(), border=border, number_format=currency_formula())
    set_cell(ws, f"C{ctx.row}", "-", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=center(), border=border)
    set_cell(ws, f"D{ctx.row}", "=K2", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=right(), border=border, number_format=currency_formula())
    set_cell(ws, f"E{ctx.row}", "=", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=center(), border=border)
    set_cell(ws, f"F{ctx.row}", "=N2", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=right(), border=border, number_format=currency_formula())
    set_cell(ws, f"G{ctx.row}", '=IF(F{0}>=0,"Net Savings","Deficit")'.format(ctx.row), font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=center(), border=border)
    style_range(ws, ctx.row, ctx.row, 8, ctx.max_col, fill_color=HEADER_BG, border=border)
    ctx.row += 1

    set_cell(ws, f"I{summary_header_row}", "Dashboard Note", font=data_font(TEXT, True), fill_color=CARD_BG, alignment=center(), border=border)
    set_cell(ws, f"I{summary_header_row+1}", "All dashboard values are linked to detailed sections below.", font=data_font(MUTED), fill_color=CARD_BG, alignment=left(), border=border)
    ws.merge_cells(start_row=summary_header_row + 1, start_column=9, end_row=summary_header_row + 7, end_column=14)
    style_range(ws, summary_header_row + 1, summary_header_row + 7, 9, 14, fill_color=CARD_BG, border=border)
    ws.freeze_panes = f"A{title_row}"


def add_fno_section(ctx: BuildContext, title: str, accent: str):
    add_section_header(ctx, title, accent)
    ws = ctx.ws
    border = thin_border()

    quarter_starts = {}
    total_rows = {}

    for sublabel in ["Options", "Futures"]:
        set_cell(ws, f"A{ctx.row}", sublabel, font=data_font(TEXT, True, 12), fill_color=CARD_BG, alignment=left(), border=border)
        style_range(ws, ctx.row, ctx.row, 1, 11, fill_color=CARD_BG, border=border)
        ctx.row += 1
        headers = ["Quarter", "Gross PnL", "Charges (Brokerage+STT+GST)", "Other Costs", "Net PnL", "Remarks", "", "", "", "", ""]
        for col_idx, text in enumerate(headers, start=1):
            if text:
                set_cell(ws, f"{get_column_letter(col_idx)}{ctx.row}", text, font=data_font(TEXT, True), fill_color=BASE_BG, alignment=center(), border=border)
            else:
                style_range(ws, ctx.row, ctx.row, col_idx, col_idx, fill_color=BASE_BG, border=border)
        ctx.row += 1
        start = ctx.row
        quarter_starts[sublabel] = start
        for q in QUARTERS:
            row_num = ctx.row
            set_cell(ws, f"A{row_num}", q, font=data_font(TEXT), fill_color=CARD_BG if (row_num - start) % 2 == 0 else BASE_BG, alignment=left(), border=border)
            for col in ["B", "C", "D"]:
                set_cell(ws, f"{col}{row_num}", 0, font=data_font(INPUT), fill_color=BASE_BG, alignment=right(), border=border, number_format=currency_formula())
            set_cell(ws, f"E{row_num}", f"=B{row_num}-C{row_num}-D{row_num}", font=data_font(TEXT, True), fill_color=CARD_BG, alignment=right(), border=border, number_format=currency_formula())
            set_cell(ws, f"F{row_num}", "", font=data_font(INPUT), fill_color=BASE_BG, alignment=left(), border=border)
            style_range(ws, row_num, row_num, 7, 11, fill_color=CARD_BG if (row_num - start) % 2 == 0 else BASE_BG, border=border)
            ctx.row += 1
        set_cell(ws, f"A{ctx.row}", f"{sublabel} Total", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=left(), border=border)
        for col in ["B", "C", "D", "E"]:
            col_letter = col
            set_cell(ws, f"{col_letter}{ctx.row}", f"=SUM({col_letter}{start}:{col_letter}{ctx.row-1})", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"F{ctx.row}", "", font=data_font(TEXT), fill_color=HEADER_BG, alignment=left(), border=border)
        style_range(ws, ctx.row, ctx.row, 7, 11, fill_color=HEADER_BG, border=border)
        total_rows[sublabel] = ctx.row
        ctx.row += 1

    option_total_row = total_rows["Options"]
    futures_total_row = total_rows["Futures"]
    option_start = quarter_starts["Options"]
    futures_start = quarter_starts["Futures"]
    set_cell(ws, f"A{ctx.row}", "Combined F&O Total", font=data_font(TEXT, True, 12), fill_color=HEADER_BG, alignment=left(), border=border)
    for q_index, col in enumerate(["G", "H", "I", "J"], start=0):
        option_q_row = option_start + q_index
        futures_q_row = futures_start + q_index
        set_cell(ws, f"{col}{ctx.row}", f"=E{option_q_row}+E{futures_q_row}", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=right(), border=border, number_format=currency_formula())
    set_cell(ws, f"K{ctx.row}", f"=E{option_total_row}+E{futures_total_row}", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=right(), border=border, number_format=currency_formula())
    ctx.row += 1


def add_equity_section(ctx: BuildContext, title: str, accent: str, holding_rows: int):
    add_section_header(ctx, title, accent)
    ws = ctx.ws
    border = thin_border()

    headers = ["Quarter", "STCG Realized", "LTCG Realized", "Charges", "Dividends", "Net PnL", "Remarks"]
    for col_idx, text in enumerate(headers, start=1):
        set_cell(ws, f"{get_column_letter(col_idx)}{ctx.row}", text, font=data_font(TEXT, True), fill_color=CARD_BG, alignment=center(), border=border)
    ctx.row += 1
    start = ctx.row
    for idx, q in enumerate(QUARTERS):
        row_num = ctx.row
        row_fill = CARD_BG if idx % 2 == 0 else BASE_BG
        set_cell(ws, f"A{row_num}", q, font=data_font(TEXT), fill_color=row_fill, alignment=left(), border=border)
        for col in ["B", "C", "D", "E"]:
            set_cell(ws, f"{col}{row_num}", 0, font=data_font(INPUT), fill_color=BASE_BG, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"F{row_num}", f"=B{row_num}+C{row_num}-D{row_num}+E{row_num}", font=data_font(TEXT, True), fill_color=row_fill, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"G{row_num}", "", font=data_font(INPUT), fill_color=BASE_BG, alignment=left(), border=border)
        ctx.row += 1
    total_row = ctx.row
    set_cell(ws, f"A{total_row}", "Total", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=left(), border=border)
    for col in ["B", "C", "D", "E", "F"]:
        set_cell(ws, f"{col}{total_row}", f"=SUM({col}{start}:{col}{total_row-1})", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=right(), border=border, number_format=currency_formula())
    set_cell(ws, f"G{total_row}", "", font=data_font(TEXT), fill_color=HEADER_BG, alignment=left(), border=border)
    ctx.row += 2

    headers = ["Stock", "Qty", "Avg Buy Price", "Current Price", "Invested", "Current Value", "Unrealized PnL"]
    for col_idx, text in enumerate(headers, start=1):
        set_cell(ws, f"{get_column_letter(col_idx)}{ctx.row}", text, font=data_font(TEXT, True), fill_color=CARD_BG, alignment=center(), border=border)
    ctx.row += 1
    holding_start = ctx.row
    for idx in range(holding_rows):
        row_num = ctx.row
        row_fill = CARD_BG if idx % 2 == 0 else BASE_BG
        set_cell(ws, f"A{row_num}", "", font=data_font(INPUT), fill_color=BASE_BG, alignment=left(), border=border)
        set_cell(ws, f"B{row_num}", 0, font=data_font(INPUT), fill_color=BASE_BG, alignment=right(), border=border)
        set_cell(ws, f"C{row_num}", 0, font=data_font(INPUT), fill_color=BASE_BG, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"D{row_num}", 0, font=data_font(INPUT), fill_color=BASE_BG, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"E{row_num}", f"=B{row_num}*C{row_num}", font=data_font(TEXT), fill_color=row_fill, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"F{row_num}", f"=B{row_num}*D{row_num}", font=data_font(TEXT), fill_color=row_fill, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"G{row_num}", f"=F{row_num}-E{row_num}", font=data_font(TEXT, True), fill_color=row_fill, alignment=right(), border=border, number_format=currency_formula())
        ctx.row += 1
    total_hold = ctx.row
    set_cell(ws, f"A{total_hold}", "Holdings Total", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=left(), border=border)
    for col in ["E", "F", "G"]:
        set_cell(ws, f"{col}{total_hold}", f"=SUM({col}{holding_start}:{col}{total_hold-1})", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=right(), border=border, number_format=currency_formula())
    for col in ["B", "C", "D"]:
        set_cell(ws, f"{col}{total_hold}", "", font=data_font(TEXT), fill_color=HEADER_BG, alignment=right(), border=border)
    ctx.row += 1


def add_crypto_section(ctx: BuildContext):
    add_section_header(ctx, "Section 6: Crypto PnL (Annual)", AMBER)
    ws = ctx.ws
    border = thin_border()
    headers = ["Category", "Gross PnL", "TDS Paid (1%)", "Net PnL", "Remarks"]
    for col_idx, text in enumerate(headers, start=1):
        set_cell(ws, f"{get_column_letter(col_idx)}{ctx.row}", text, font=data_font(TEXT, True), fill_color=CARD_BG, alignment=center(), border=border)
    ctx.row += 1
    start = ctx.row
    for idx, label in enumerate(["Spot Trading", "Derivatives/Futures", "Airdrops/Staking"]):
        row_num = ctx.row
        row_fill = CARD_BG if idx % 2 == 0 else BASE_BG
        set_cell(ws, f"A{row_num}", label, font=data_font(TEXT), fill_color=row_fill, alignment=left(), border=border)
        set_cell(ws, f"B{row_num}", 0, font=data_font(INPUT), fill_color=BASE_BG, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"C{row_num}", 0, font=data_font(INPUT), fill_color=BASE_BG, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"D{row_num}", f"=B{row_num}-C{row_num}", font=data_font(TEXT, True), fill_color=row_fill, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"E{row_num}", "", font=data_font(INPUT), fill_color=BASE_BG, alignment=left(), border=border)
        ctx.row += 1
    total_row = ctx.row
    set_cell(ws, f"A{total_row}", "Total", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=left(), border=border)
    for col in ["B", "C", "D"]:
        set_cell(ws, f"{col}{total_row}", f"=SUM({col}{start}:{col}{total_row-1})", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=right(), border=border, number_format=currency_formula())
    set_cell(ws, f"E{total_row}", "", font=data_font(TEXT), fill_color=HEADER_BG, alignment=left(), border=border)
    ctx.row += 1
    ws.merge_cells(start_row=ctx.row, start_column=1, end_row=ctx.row, end_column=ctx.max_col)
    set_cell(ws, f"A{ctx.row}", "Tax Note: Section 115BBH applies. Crypto gains attract 30% tax plus cess, losses cannot be set off, and 1% TDS may apply subject to law/thresholds.", font=data_font(AMBER, False, 10), fill_color=BASE_BG, alignment=left(), border=border)
    ctx.row += 2

    headers = ["Token", "Qty", "Avg Buy (INR)", "Current Price (INR)", "Unrealized PnL"]
    for col_idx, text in enumerate(headers, start=1):
        set_cell(ws, f"{get_column_letter(col_idx)}{ctx.row}", text, font=data_font(TEXT, True), fill_color=CARD_BG, alignment=center(), border=border)
    ctx.row += 1
    start_hold = ctx.row
    for idx in range(5):
        row_num = ctx.row
        row_fill = CARD_BG if idx % 2 == 0 else BASE_BG
        set_cell(ws, f"A{row_num}", "", font=data_font(INPUT), fill_color=BASE_BG, alignment=left(), border=border)
        set_cell(ws, f"B{row_num}", 0, font=data_font(INPUT), fill_color=BASE_BG, alignment=right(), border=border)
        set_cell(ws, f"C{row_num}", 0, font=data_font(INPUT), fill_color=BASE_BG, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"D{row_num}", 0, font=data_font(INPUT), fill_color=BASE_BG, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"E{row_num}", f"=(D{row_num}-C{row_num})*B{row_num}", font=data_font(TEXT, True), fill_color=row_fill, alignment=right(), border=border, number_format=currency_formula())
        ctx.row += 1
    set_cell(ws, f"A{ctx.row}", "Holdings Total", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=left(), border=border)
    set_cell(ws, f"E{ctx.row}", f"=SUM(E{start_hold}:E{ctx.row-1})", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=right(), border=border, number_format=currency_formula())
    style_range(ws, ctx.row, ctx.row, 2, 4, fill_color=HEADER_BG, border=border)
    ctx.row += 1


def add_business_section(ctx: BuildContext):
    add_section_header(ctx, "Section 7: Business / Freelance Income", CYAN)
    ws = ctx.ws
    border = thin_border()
    headers = ["Month", "Client/Source", "Revenue", "Expenses", "Net Profit (Revenue - Expenses)", "Notes"]
    for col_idx, text in enumerate(headers, start=1):
        set_cell(ws, f"{get_column_letter(col_idx)}{ctx.row}", text, font=data_font(TEXT, True), fill_color=CARD_BG, alignment=center(), border=border)
    ctx.row += 1
    start = ctx.row
    for idx, month in enumerate(MONTHS_FULL):
        row_num = ctx.row
        row_fill = CARD_BG if idx % 2 == 0 else BASE_BG
        set_cell(ws, f"A{row_num}", month, font=data_font(TEXT), fill_color=row_fill, alignment=left(), border=border)
        set_cell(ws, f"B{row_num}", "", font=data_font(INPUT), fill_color=BASE_BG, alignment=left(), border=border)
        set_cell(ws, f"C{row_num}", 0, font=data_font(INPUT), fill_color=BASE_BG, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"D{row_num}", 0, font=data_font(INPUT), fill_color=BASE_BG, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"E{row_num}", f"=C{row_num}-D{row_num}", font=data_font(TEXT, True), fill_color=row_fill, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"F{row_num}", "", font=data_font(INPUT), fill_color=BASE_BG, alignment=left(), border=border)
        ctx.row += 1
    total_row = ctx.row
    set_cell(ws, f"A{total_row}", "FY Total", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=left(), border=border)
    for col in ["C", "D", "E"]:
        set_cell(ws, f"{col}{total_row}", f"=SUM({col}{start}:{col}{total_row-1})", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=right(), border=border, number_format=currency_formula())
    set_cell(ws, f"B{total_row}", "", font=data_font(TEXT), fill_color=HEADER_BG, alignment=left(), border=border)
    set_cell(ws, f"F{total_row}", "", font=data_font(TEXT), fill_color=HEADER_BG, alignment=left(), border=border)
    ctx.row += 2

    headers = ["Quarter", "Revenue", "Expenses", "Net Profit"]
    for offset, text in enumerate(headers):
        set_cell(ws, f"{get_column_letter(11 + offset)}{ctx.row}", text, font=data_font(TEXT, True), fill_color=CARD_BG, alignment=center(), border=border)
    ctx.row += 1
    q_ranges = [(start, start + 2), (start + 3, start + 5), (start + 6, start + 8), (start + 9, start + 11)]
    q_start_row = ctx.row
    for idx, (label, rng) in enumerate(zip(QUARTERS, q_ranges)):
        row_num = ctx.row
        row_fill = CARD_BG if idx % 2 == 0 else BASE_BG
        r1, r2 = rng
        set_cell(ws, f"K{row_num}", label, font=data_font(TEXT), fill_color=row_fill, alignment=left(), border=border)
        set_cell(ws, f"L{row_num}", f"=SUM(C{r1}:C{r2})", font=data_font(TEXT), fill_color=row_fill, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"M{row_num}", f"=SUM(D{r1}:D{r2})", font=data_font(TEXT), fill_color=row_fill, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"N{row_num}", f"=SUM(E{r1}:E{r2})", font=data_font(TEXT, True), fill_color=row_fill, alignment=right(), border=border, number_format=currency_formula())
        ctx.row += 1
    return q_start_row


def add_mf_section(ctx: BuildContext, title: str, accent: str):
    add_section_header(ctx, title, accent)
    ws = ctx.ws
    border = thin_border()
    headers = ["Fund Name", "Type", "Total Invested", "Monthly SIP", "Current Value", "Abs Return", "% Return", "Remarks"]
    for col_idx, text in enumerate(headers, start=1):
        set_cell(ws, f"{get_column_letter(col_idx)}{ctx.row}", text, font=data_font(TEXT, True), fill_color=CARD_BG, alignment=center(), border=border)
    ctx.row += 1
    start = ctx.row
    for idx in range(8):
        row_num = ctx.row
        row_fill = CARD_BG if idx % 2 == 0 else BASE_BG
        set_cell(ws, f"A{row_num}", "", font=data_font(INPUT), fill_color=BASE_BG, alignment=left(), border=border)
        set_cell(ws, f"B{row_num}", "", font=data_font(INPUT), fill_color=BASE_BG, alignment=left(), border=border)
        for col in ["C", "D", "E"]:
            set_cell(ws, f"{col}{row_num}", 0, font=data_font(INPUT), fill_color=BASE_BG, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"F{row_num}", f"=E{row_num}-C{row_num}", font=data_font(TEXT), fill_color=row_fill, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"G{row_num}", f"=IFERROR(F{row_num}/C{row_num},0)", font=data_font(TEXT), fill_color=row_fill, alignment=right(), border=border, number_format=percent_formula())
        set_cell(ws, f"H{row_num}", "", font=data_font(INPUT), fill_color=BASE_BG, alignment=left(), border=border)
        ctx.row += 1
    set_cell(ws, f"A{ctx.row}", "Portfolio Total", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=left(), border=border)
    for col in ["C", "D", "E", "F"]:
        set_cell(ws, f"{col}{ctx.row}", f"=SUM({col}{start}:{col}{ctx.row-1})", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=right(), border=border, number_format=currency_formula())
    set_cell(ws, f"G{ctx.row}", f"=IFERROR(F{ctx.row}/C{ctx.row},0)", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=right(), border=border, number_format=percent_formula())
    set_cell(ws, f"H{ctx.row}", "", font=data_font(TEXT), fill_color=HEADER_BG, alignment=left(), border=border)
    ctx.row += 1


def add_expenses_section(ctx: BuildContext):
    add_section_header(ctx, "Section 10: Personal Expenses", LOSS)
    ws = ctx.ws
    border = thin_border()
    headers = ["Category", *MONTHS_FULL, "FY Total"]
    for col_idx, text in enumerate(headers, start=1):
        set_cell(ws, f"{get_column_letter(col_idx)}{ctx.row}", text, font=data_font(TEXT, True), fill_color=CARD_BG, alignment=center(), border=border)
    ctx.row += 1
    start = ctx.row
    for idx, cat in enumerate(EXPENSE_CATS):
        row_num = ctx.row
        row_fill = CARD_BG if idx % 2 == 0 else BASE_BG
        set_cell(ws, f"A{row_num}", cat, font=data_font(TEXT), fill_color=row_fill, alignment=left(), border=border)
        for col_idx in range(2, 14):
            set_cell(ws, f"{get_column_letter(col_idx)}{row_num}", 0, font=data_font(INPUT), fill_color=BASE_BG, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"N{row_num}", f"=SUM(B{row_num}:M{row_num})", font=data_font(TEXT, True), fill_color=row_fill, alignment=right(), border=border, number_format=currency_formula())
        ctx.row += 1
    monthly_total_row = ctx.row
    set_cell(ws, f"A{monthly_total_row}", "Monthly Total", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=left(), border=border)
    for col_idx in range(2, 14):
        col_letter = get_column_letter(col_idx)
        set_cell(ws, f"{col_letter}{monthly_total_row}", f"=SUM({col_letter}{start}:{col_letter}{monthly_total_row-1})", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=right(), border=border, number_format=currency_formula())
    set_cell(ws, f"N{monthly_total_row}", f"=SUM(N{start}:N{monthly_total_row-1})", font=data_font(TEXT, True), fill_color=HEADER_BG, alignment=right(), border=border, number_format=currency_formula())
    ctx.row += 2

    headers = ["Category", "FY Total", "Monthly Avg", "% of Total"]
    for offset, text in enumerate(headers):
        set_cell(ws, f"{get_column_letter(1 + offset)}{ctx.row}", text, font=data_font(TEXT, True), fill_color=CARD_BG, alignment=center(), border=border)
    ctx.row += 1
    for idx, cat in enumerate(EXPENSE_CATS):
        row_num = ctx.row
        src_row = start + idx
        row_fill = CARD_BG if idx % 2 == 0 else BASE_BG
        set_cell(ws, f"A{row_num}", cat, font=data_font(TEXT), fill_color=row_fill, alignment=left(), border=border)
        set_cell(ws, f"B{row_num}", f"=N{src_row}", font=data_font(TEXT), fill_color=row_fill, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"C{row_num}", f"=IFERROR(B{row_num}/12,0)", font=data_font(TEXT), fill_color=row_fill, alignment=right(), border=border, number_format=currency_formula())
        set_cell(ws, f"D{row_num}", f"=IFERROR(B{row_num}/$N${monthly_total_row},0)", font=data_font(TEXT), fill_color=row_fill, alignment=right(), border=border, number_format=percent_formula())
        ctx.row += 1


def apply_sheet_layout(ctx: BuildContext):
    ws = ctx.ws
    content_end_row = ctx.row - 1
    widths = {
        "A": 24,
        "B": 16,
        "C": 16,
        "D": 16,
        "E": 18,
        "F": 18,
        "G": 18,
        "H": 18,
        "I": 14,
        "J": 14,
        "K": 16,
        "L": 16,
        "M": 16,
        "N": 16,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.3, bottom=0.3, header=0.1, footer=0.1)
    ws.print_area = f"A1:N{content_end_row}"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.auto_filter.ref = f"A3:G9"


def apply_conditional_formatting(ws):
    positive_ranges = [
        "E14:E22", "E27:E35", "F41:F45", "G48:G56", "F59:F63", "G66:G72", "D78:D81",
        "E87:E99", "N91:N94", "F103:F111", "G103:G111", "F116:F124", "G116:G124",
        "N131:N141", "B144:D153",
    ]
    for cell_range in positive_ranges:
        ws.conditional_formatting.add(cell_range, CellIsRule(operator="greaterThan", formula=["0"], font=Font(color=PROFIT)))
        ws.conditional_formatting.add(cell_range, CellIsRule(operator="lessThan", formula=["0"], font=Font(color=LOSS)))


def build_workbook(output_path: Path) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "FY26 Tracker"
    ctx = BuildContext(wb=wb, ws=ws)
    ws.sheet_properties.tabColor = HEADER_BG
    style_range(ws, 1, 200, 1, 14, fill_color=BASE_BG, font=data_font(TEXT), alignment=left())

    add_dashboard(ctx)
    add_spacer(ctx)
    add_fno_section(ctx, "Section 2: F&O PnL — Ayush", HEADER_BG)
    add_spacer(ctx)
    add_fno_section(ctx, "Section 3: F&O PnL — Family", PURPLE)
    add_spacer(ctx)
    add_equity_section(ctx, "Section 4: Equity PnL — Ayush", PROFIT, 8)
    add_spacer(ctx)
    add_equity_section(ctx, "Section 5: Equity PnL — Family", PURPLE, 6)
    add_spacer(ctx)
    add_crypto_section(ctx)
    add_spacer(ctx)
    add_business_section(ctx)
    add_spacer(ctx)
    add_mf_section(ctx, "Section 8: Mutual Funds — Ayush", HEADER_BG)
    add_spacer(ctx)
    add_mf_section(ctx, "Section 9: Mutual Funds — Family", PURPLE)
    add_spacer(ctx)
    add_expenses_section(ctx)
    apply_sheet_layout(ctx)
    apply_conditional_formatting(ws)
    wb.save(output_path)
    return wb


def build_pdf(output_path: Path, workbook_path: Path):
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(A3),
        leftMargin=0.35 * inch,
        rightMargin=0.35 * inch,
        topMargin=0.35 * inch,
        bottomMargin=0.35 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleDark",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=hex_color(TEXT),
        alignment=1,
        spaceAfter=8,
    )
    body_style = ParagraphStyle(
        "BodyDark",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        textColor=hex_color(TEXT),
        leading=12,
        spaceAfter=6,
    )
    section_style = ParagraphStyle(
        "SectionDark",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=11,
        textColor=hex_color(TEXT),
        backColor=f"#{HEADER_BG}",
        borderPadding=6,
        spaceBefore=4,
        spaceAfter=6,
    )

    story = [
        Paragraph("FY 2025-26 | Consolidated Financial Tracker", title_style),
        Paragraph("Prepared for Ayush. This PDF accompanies the Excel workbook and preserves the full single-sheet tracker structure, dark-theme design cues, and section coverage.", body_style),
        Paragraph(f"Workbook file: {workbook_path.name}", body_style),
        Spacer(1, 0.12 * inch),
    ]

    sections = [
        ("Dashboard / Consolidated Summary", [
            ["User Name", "Ayush", "Prepared On", "Workbook formula date", "Total Income", "Linked", "Total Expenses", "Linked", "Net Savings", "Linked"],
            ["Income Stream", "Q1", "Q2", "Q3", "Q4", "FY Total", "% of Total", "", "", ""],
            ["F&O Trading (You)", "Linked", "Linked", "Linked", "Linked", "Linked", "Formula", "", "", ""],
            ["F&O Trading (Family)", "Linked", "Linked", "Linked", "Linked", "Linked", "Formula", "", "", ""],
            ["Equity (You)", "Linked", "Linked", "Linked", "Linked", "Linked", "Formula", "", "", ""],
            ["Equity (Family)", "Linked", "Linked", "Linked", "Linked", "Linked", "Formula", "", "", ""],
            ["Crypto", "Linked", "0", "0", "0", "Linked", "Formula", "", "", ""],
            ["Business/Freelance", "Linked", "Linked", "Linked", "Linked", "Linked", "Formula", "", "", ""],
            ["Net Position", "Total Income", "-", "Total Expenses", "=", "Net Savings / Deficit", "", "", "", ""],
        ], PROFIT),
        ("F&O PnL - Ayush", [
            ["Type", "Quarter", "Gross PnL", "Charges", "Other Costs", "Net PnL", "Remarks"],
            ["Options", "Q1 (Apr-Jun)", "", "", "", "Formula", ""],
            ["Options", "Q2 (Jul-Sep)", "", "", "", "Formula", ""],
            ["Options", "Q3 (Oct-Dec)", "", "", "", "Formula", ""],
            ["Options", "Q4 (Jan-Mar)", "", "", "", "Formula", ""],
            ["Options", "Total", "", "", "", "Formula", ""],
            ["Futures", "Q1 (Apr-Jun)", "", "", "", "Formula", ""],
            ["Futures", "Q2 (Jul-Sep)", "", "", "", "Formula", ""],
            ["Futures", "Q3 (Oct-Dec)", "", "", "", "Formula", ""],
            ["Futures", "Q4 (Jan-Mar)", "", "", "", "Formula", ""],
            ["Futures", "Total", "", "", "", "Formula", ""],
            ["Combined", "FY Total", "", "", "", "Formula", ""],
        ], HEADER_BG),
        ("F&O PnL - Family", [
            ["Type", "Quarter", "Gross PnL", "Charges", "Other Costs", "Net PnL", "Remarks"],
            ["Options", "Q1 (Apr-Jun)", "", "", "", "Formula", ""],
            ["Options", "Q2 (Jul-Sep)", "", "", "", "Formula", ""],
            ["Options", "Q3 (Oct-Dec)", "", "", "", "Formula", ""],
            ["Options", "Q4 (Jan-Mar)", "", "", "", "Formula", ""],
            ["Options", "Total", "", "", "", "Formula", ""],
            ["Futures", "Q1 (Apr-Jun)", "", "", "", "Formula", ""],
            ["Futures", "Q2 (Jul-Sep)", "", "", "", "Formula", ""],
            ["Futures", "Q3 (Oct-Dec)", "", "", "", "Formula", ""],
            ["Futures", "Q4 (Jan-Mar)", "", "", "", "Formula", ""],
            ["Futures", "Total", "", "", "", "Formula", ""],
            ["Combined", "FY Total", "", "", "", "Formula", ""],
        ], PURPLE),
        ("Equity PnL - Ayush", [
            ["Quarter", "STCG", "LTCG", "Charges", "Dividends", "Net PnL", "Remarks"],
            *[[q, "", "", "", "", "Formula", ""] for q in QUARTERS],
            ["Total", "", "", "", "", "Formula", ""],
            ["", "", "", "", "", "", ""],
            ["Current Holdings", "Qty", "Avg Buy Price", "Current Price", "Invested", "Current Value", "Unrealized PnL"],
            *[[f"Holding {i}", "", "", "", "Formula", "Formula", "Formula"] for i in range(1, 9)],
            ["Holdings Total", "", "", "", "Formula", "Formula", "Formula"],
        ], PROFIT),
        ("Equity PnL - Family", [
            ["Quarter", "STCG", "LTCG", "Charges", "Dividends", "Net PnL", "Remarks"],
            *[[q, "", "", "", "", "Formula", ""] for q in QUARTERS],
            ["Total", "", "", "", "", "Formula", ""],
            ["", "", "", "", "", "", ""],
            ["Current Holdings", "Qty", "Avg Buy Price", "Current Price", "Invested", "Current Value", "Unrealized PnL"],
            *[[f"Holding {i}", "", "", "", "Formula", "Formula", "Formula"] for i in range(1, 7)],
            ["Holdings Total", "", "", "", "Formula", "Formula", "Formula"],
        ], PURPLE),
        ("Crypto PnL (Annual)", [
            ["Category", "Gross PnL", "TDS Paid (1%)", "Net PnL", "Remarks"],
            ["Spot Trading", "", "", "Formula", ""],
            ["Derivatives/Futures", "", "", "Formula", ""],
            ["Airdrops/Staking", "", "", "Formula", ""],
            ["Total", "", "", "Formula", ""],
            ["Tax Note", "Section 115BBH: 30% flat, no set-off, 1% TDS subject to law", "", "", ""],
            ["", "", "", "", ""],
            ["Current Holdings", "Qty", "Avg Buy (INR)", "Current Price (INR)", "Unrealized PnL"],
            *[[f"Token {i}", "", "", "", "Formula"] for i in range(1, 6)],
            ["Holdings Total", "", "", "", "Formula"],
        ], AMBER),
        ("Business / Freelance Income", [
            ["Month", "Client/Source", "Revenue", "Expenses", "Net Profit", "Notes"],
            *[[month, "", "", "", "Formula", ""] for month in MONTHS_FULL],
            ["FY Total", "", "Formula", "Formula", "Formula", ""],
            ["", "", "", "", "", ""],
            ["Quarter", "Revenue", "Expenses", "Net Profit", "", ""],
            *[[q, "Formula", "Formula", "Formula", "", ""] for q in QUARTERS],
        ], CYAN),
        ("Mutual Funds - Ayush", [
            ["Fund Name", "Type", "Total Invested", "Monthly SIP", "Current Value", "Abs Return", "% Return", "Remarks"],
            *[[f"Fund {i}", "", "", "", "", "Formula", "Formula", ""] for i in range(1, 9)],
            ["Portfolio Total", "", "Formula", "Formula", "Formula", "Formula", "Formula", ""],
        ], HEADER_BG),
        ("Mutual Funds - Family", [
            ["Fund Name", "Type", "Total Invested", "Monthly SIP", "Current Value", "Abs Return", "% Return", "Remarks"],
            *[[f"Fund {i}", "", "", "", "", "Formula", "Formula", ""] for i in range(1, 9)],
            ["Portfolio Total", "", "Formula", "Formula", "Formula", "Formula", "Formula", ""],
        ], PURPLE),
        ("Personal Expenses", [
            ["Category", *MONTHS_FULL, "FY Total"],
            *[[cat, *["" for _ in MONTHS_FULL], "Formula"] for cat in EXPENSE_CATS],
            ["Monthly Total", *["Formula" for _ in MONTHS_FULL], "Formula"],
            ["", *["" for _ in MONTHS_FULL], ""],
            ["Category", "FY Total", "Monthly Avg", "% of Total", "", "", "", "", "", "", "", "", "", ""],
            *[[cat, "Formula", "Formula", "Formula", "", "", "", "", "", "", "", "", "", ""] for cat in EXPENSE_CATS],
        ], LOSS),
    ]

    for idx, (title, data, accent) in enumerate(sections):
        story.append(Paragraph(title, ParagraphStyle(
            f"Section_{idx}",
            parent=section_style,
            backColor=f"#{accent}",
        )))
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), hex_color(CARD_BG)),
            ("TEXTCOLOR", (0, 0), (-1, -1), hex_color(TEXT)),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 1), (-1, -1), hex_color(BASE_BG)),
            ("GRID", (0, 0), (-1, -1), 0.4, hex_color(BORDER)),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [hex_color(CARD_BG), hex_color(BASE_BG)]),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(table)
        story.append(Spacer(1, 0.15 * inch))
        if idx < len(sections) - 1:
            story.append(PageBreak())

    story.append(PageBreak())
    story.append(Paragraph("Notes", section_style))
    notes = [
        "The Excel workbook is the primary editable tracker and contains the live formulas, conditional formatting, and dashboard references.",
        "Blue-font cells are intended for user input. Profit-oriented formula cells are conditionally formatted green; loss-making values are red.",
        "The workbook includes a user-name field and is prepared as a single-sheet layout from April 2025 through March 2026.",
        "The PDF mirrors the full tracker structure section by section; for calculations and data entry, use the .xlsx file.",
    ]
    for note in notes:
        story.append(Paragraph(f"- {note}", body_style))

    def paint_bg(canvas, _doc):
        canvas.saveState()
        canvas.setFillColor(hex_color(SPACER_BG))
        canvas.rect(0, 0, _doc.pagesize[0], _doc.pagesize[1], fill=1, stroke=0)
        canvas.restoreState()

    doc.build(story, onFirstPage=paint_bg, onLaterPages=paint_bg)


def main():
    out_dir = Path(__file__).resolve().parent
    xlsx_path = out_dir / "FY26_Consolidated_Financial_Tracker.xlsx"
    pdf_path = out_dir / "FY26_Consolidated_Financial_Tracker.pdf"
    build_workbook(xlsx_path)
    build_pdf(pdf_path, xlsx_path)
    print(f"Created {xlsx_path}")
    print(f"Created {pdf_path}")


if __name__ == "__main__":
    main()
